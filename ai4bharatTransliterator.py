from ai4bharat.transliteration import XlitEngine
import openpyxl
import argparse
import sys
import io

e = XlitEngine("hi", beam_width=10)

wb = openpyxl.load_workbook('input_tweets.xlsx')
sheet = wb['Sheet1']

new_wb = openpyxl.Workbook()
new_sheet = new_wb.active
new_sheet.title = 'Hindi_tweets'

if __name__ == '__main__':
    prog = 'tranliterator'
    description = 'Transliterate tweets'
    parser = argparse.ArgumentParser(prog=prog,
                                     description=description)

    parser.add_argument('-o',
                        '--output',
                        metavar='',
                        dest='outfile',
                        type=str,
                        help='<output-file>')
    args = parser.parse_args(sys.argv[1:])
    ofp = io.open(args.outfile, mode='w', encoding='utf-8')

    for row in sheet.iter_rows(values_only=True):
        tweet = row[0]
        try:
            hindi_tweet = e.translit_sentence(tweet)
            print(hindi_tweet)
            ofp.write('%s\n' % hindi_tweet)
        except Exception as ex:
            print(f'Failed to transliterate tweet "{tweet}" - {ex}')

    ofp.close()